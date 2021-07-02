import sys
import openpyxl
import numpy as np

class Instancia:
    def __init__(self, dd = None, tp = None):
        self.due_dates = dd
        self.tiempos_procesamiento = tp
    #fed
    def __str__(self):
        return "{Due Dates: %s, \nTiempos de procesamiento:\n %s\n}" % (self.due_dates, np.array(self.tiempos_procesamiento))
    #fed
    def __repr__(self):
        return str(self)
    #fed
#ssalc

def read_data_XLSX():
    ## --------------------- READ DATA EXCEL WHIT OPENPYXL -----------------------------------
    
    ## OPEN FILE AND SAVE IN VARIABLE (archivo)
    archivo = openpyxl.load_workbook("Instancias de prueba.xlsx", data_only = True)

    INSTANCIAS = []
    
    ## READ SHEETS OF FILE AND SELECT THE THE FIRST SHEET
    sheets = archivo.sheetnames
    
    for s in sheets:
        sheet = archivo[s]
        
        ## Número de máquinas y de trabajos
        num_trabajos = sheet['C4'].value
        num_maquinas = sheet['C5'].value
        
        ## Obtener due dates
        due_dates = []
        for i in range(num_trabajos):
            due_dates.append(int(sheet.cell(row = 9 + i, column = 3).value))
        
        
        ## Obtener tiempos de procesos
        TP = []
        for i in range(num_trabajos):
            t = []
            for j in range(num_maquinas):
                t.append(int(sheet.cell(row = 5 + i, column = 6 + j).value))
            
            TP.append(t)
        
        
        
        INSTANCIAS.append(Instancia(due_dates,TP))
    
    #print(np.array(INSTANCIAS))
    return(INSTANCIAS)
#fed

def principal(argv):
    read_data_XLSX()
#fed



if __name__ == "__main__":
    principal(sys.argv)
