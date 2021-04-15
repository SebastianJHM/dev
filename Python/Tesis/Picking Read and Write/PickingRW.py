from pyomo.environ import *
from pyomo.opt import SolverFactory
import openpyxl
from openpyxl.utils import get_column_letter
import copy
import random
import xlsxwriter
import sys
import numpy as np

def lecturaDatos():
    ## --------------------- LEER ARCHIVO EXCEL CON OPENPYXL -----------------------------------
    ## Abrir el archivo y guardar en la variable archivo
    archivo = openpyxl.load_workbook('Datos_Picking.xlsx', data_only = True)

    ## Leer las hojas del archivo, Seleccionar la primera hoja y la columna D
    sheets = archivo.sheetnames
    sheetDatos = archivo[sheets[0]]

    ## NODOS
    n = []
    columnA = sheetDatos['A']
    for x in range(len(columnA)): 
        if (type(columnA[x].value) != str and columnA[x].value != None):
            n.append(columnA[x].value)
        #fi
    #rof

    ## O
    o = []
    columnB = sheetDatos['B']
    for x in range(len(columnB)): 
        if (type(columnB[x].value) != str and columnB[x].value != None):
            o.append(columnB[x].value)
        #fi
    #rof


    ## REF
    ref = []
    columnC = sheetDatos['C']
    for x in range(len(columnC)): 
        if (type(columnC[x].value) != str and columnC[x].value != None):
            ref.append(columnC[x].value)
        #fi
    #rof

    ## NOD_REF
    nr = []
    columnF = sheetDatos['F']
    for x in range(len(columnF)): 
        if (type(columnF[x].value) != str and columnF[x].value != None):
            nr.append(columnF[x].value)
        #fi
    #rof
    ubic = dict(zip(ref, nr))


    ## ORDENES y R
    lect = []
    for i in range(len(o)):
        column = sheetDatos[get_column_letter( 8 + i )]
        orden = []
        for x in range(len(column)): 
            if (type(column[x].value) != str and column[x].value != None):
                orden.append(column[x].value)
            #fi
        #rof
        lect.append(orden)
    #rof
    ordenesExcel = dict(zip(o, lect))

    lectR = copy.deepcopy(lect) ## Esto es demencial
    for x in lectR:
        x.remove(0)
    #rof
    rExcel = dict(zip(o, lectR))


    ## Distancia
    dist = []
    sheetDistancias = archivo[sheets[1]]
    for i in  range(sheetDistancias.min_row + 1,sheetDistancias.max_row + 1):
        y = []
        for j in range(sheetDistancias.min_column + 1, sheetDistancias.max_column + 1):
            y.append(float(sheetDistancias.cell(row = i, column = j).value))
        #rof
        dist.append(y)
    #rof
    ## -------------------------------------------------------------------------------
    
    return(n, o, ref, ubic, ordenesExcel, rExcel, dist)
#fed

def modeloLineal(n, o, ref, ubic, ordenesExcel, rExcel, dist):
    

    

    ## Conjuntos ------------------------------------------------------------------------
    model.NODOS = Set( initialize = n )
    model.O = Set( initialize = o )
    model.REF = Set( initialize = ref )
    def setORDENES(model, i):
        return(list(ordenesExcel[i]))
    #fed
    model.ORDENES = Set(model.O, initialize = setORDENES)
    def setR(model, i):
        return(list(rExcel[i]))
    #fed
    model.R = Set(model.O, initialize = setR)



    def junte(model):
        return((o,i,j) for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o] )
    #fed
    model.OROR = Set(dimen =3, initialize = junte )

    def junte2(model):
        return((o,i) for o in model.O for i in model.ORDENES[o] )
    #fed
    model.OX = Set(dimen = 2, initialize = junte2 )

    def junte3(model):
        return((o,i,j) for o in model.O for i in model.R[o] for j in model.R[o] )
    #fed
    model.ORR = Set(dimen = 3, initialize = junte3 )

    def junte4(model):
        return((o,i) for o in model.O for i in model.R[o] )
    #fed
    model.OR  = Set(dimen = 2, initialize = junte4 )


    ## Parámetros -------------------------------------------------------------------------------
    def paramDistancia(model, i, j):
        return(dist[i][j])
    #fed
    model.Distancia = Param(model.NODOS,model.NODOS, initialize = paramDistancia)
    def paramNodRef(model, i):
        return(ubic[i])
    #fed
    model.Nod_Ref = Param(model.REF, initialize = paramNodRef) 

    def dinit(model, i ,j):
        return model.Distancia[model.Nod_Ref[i],model.Nod_Ref[j]]
    #fed
    model.Distancia_R = Param(model.REF,model.REF, initialize = dinit,mutable = True)

    ## Variables ---------------------------------------------------------------------------------
    model.x = Var(model.OROR, domain = Binary)
    model.aux = Var(model.OR)
    model.Dist_ORD = Var(model.O)


    ## Función Objetivo ---------------------------------------------------------------------------------
    def ObjFunc(model):
        return sum(model.Distancia_R[i,j]*model.x[o,i,j] for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o])
    #fed
    model.FO = Objective(rule = ObjFunc)


    ## Restricciones ---------------------------------------------------------------------------------
    def r1(model, o, i):
        return sum( model.x[o,i,j] for j in model.ORDENES[o]) == 1
    #fed
    model.r1 = Constraint( model.OX, rule = r1 )

    def r2(model, o, j):
        return sum( model.x[o,i,j] for i in model.ORDENES[o]) == 1
    #fed
    model.r2 = Constraint( model.OX,  rule = r2 )

    def r3(model, o, i):
        return model.x[o,i,i] == 0
    #fed
    model.r3 = Constraint( model.OX,  rule = r3 )

    def r4(model, o, i, j):
        if ( i != j ):
            return model.aux[o,i] - model.aux[o,j] + len(model.R[o])*model.x[o,i,j] <= len(model.R[o]) - 1 
        return Constraint.Skip
    #fed
    model.r4 = Constraint( model.ORR,  rule = r4 )

    def r5(model, o):
        return model.Dist_ORD[o] == sum( model.x[o,i,j]*model.Distancia_R[i,j] for i in model.ORDENES[o] for j in model.ORDENES[o] )
    #fed
    model.r5 = Constraint( model.O, rule = r5)


#fed


def imprimirResultadosXLSX():
    ## --------------------- GUARDAR LOS RESULTADOS EN ARCHIVO EXCEL -----------------------------
    
    # Crear libro y añadir un hoja
    workbook = xlsxwriter.Workbook('Resultados.xlsx')

    ## Formatos
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center'})
        
    ## Crear hoja
    worksheet = workbook.add_worksheet('Resultados')
    worksheet.set_column(0, 0, 20)

    #worksheet.merge_range('A1:D1', "SOLUCIÓN DEL EJERCICIO", merge_format)
    worksheet.merge_range(0, 0, 0, 3, "SOLUCIÓN DEL EJERCICIO", merge_format)

    worksheet.write(1, 0, "Distancia Total: ", merge_format)
    worksheet.write(1, 1, value(instance.FO), cell_format)

    row=3
    col=0
    for o in instance.O:
        worksheet.merge_range(row, col, row, 3, "Orden "+ str(o), merge_format)
        row += 1

        worksheet.write(row, col,"Distancia Recorrida: ", cell_format)
        worksheet.write(row, col + 1,instance.Dist_ORD[o].value, cell_format)
        row += 1
        
        worksheet.write(row, col,"Ruta: ", cell_format)
        row += 1
        
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    worksheet.write(row, col + 0,"Del nodo ", cell_format)
                    worksheet.write(row, col + 1,i, cell_format)
                    worksheet.write(row, col + 2," al nodo", cell_format)
                    worksheet.write(row, col + 3,j, cell_format)
                #fi
            #rof
            row += 1
        #rof
        row += 1
    #rof
        
        
    # Hay que incluir workbook.close()
    workbook.close()
    ##  -------------------------------------------------------------------------------------
#fed




## Impresión de Resultados
def imprimirResultadosConsola():
    print("\n\n")
    print("SOLUCIÓN DEL EJERCICIO")
    print("--------------------------")
    print("\n")
    print("Distancia Total: ",value(instance.FO))
    print("\n")
    for o in instance.O:
        print("--------- Orden ",o,"----------")
        print("Distancia Recorrida: ",instance.Dist_ORD[o].value)
        print("Ruta: ")
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    print("Del nodo ",i," al nodo",j)
                #fi
            #rof
        #rof
    #rof
#fed





## Lectura de datos con OPENPYXL
n, o, ref, ubic, ordenesExcel, rExcel, dist = lecturaDatos()


## Crear Modelo Lineal con PYOMO, resolverlo con GLPK y guardarlo en variable model
opt = SolverFactory('cplex', executable="C:\\Program Files\\IBM\\ILOG\\CPLEX_Studio1210\\cplex\\bin\\x64_win64\\cplex")
#opt = SolverFactory('glpk')
model = AbstractModel()
modeloLineal(n, o, ref, ubic, ordenesExcel, rExcel, dist)

## Crear una instancia del modelo y ejecutarla
instance = model.create_instance() 
opt.options['timelimit'] = 120
results = opt.solve(instance, tee = False)
#instance.display()


## Imprimir resultado por consola
imprimirResultadosConsola()

## Crear archivo e imprimir resultados xlsxwriter
imprimirResultadosXLSX()


