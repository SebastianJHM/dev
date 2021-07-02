import openpyxl
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import xlsxwriter
import sys

def read_data_XLSX_Alloc():
    ## --------------------- READ DATA EXCEL WHIT OPENPYXL -----------------------------------
    
    ## OPEN FILE AND SAVE IN VARIABLE (archivo)
    archivo = openpyxl.load_workbook("Data_Allocation.xlsx", data_only = True)

    ## READ SHEETS OF FILE AND SELECT THE THE FIRST SHEET
    sheets = archivo.sheetnames
    sheetDatos = archivo[sheets[0]]
    
    ## REFERENCES (set_REF)
    set_REF = []
    column = sheetDatos['B']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_REF.append(column[x].value)
        
    
    
    ## BOXES WIDTH (param_AnchoCaja)
    lect = []
    column = sheetDatos['C']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_AnchoCaja = dict(zip(set_REF, lect))
    
    ## IF REFERENCE IS IN ROWS (param_TipoDist)
    lect = []
    column = sheetDatos['D']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_TipoDist = dict(zip(set_REF, lect))
    
    ## NUMBERS SPACES OF REFERENCE (param_Espacios)
    lect = []
    column = sheetDatos['E']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_Espacios = dict(zip(set_REF, lect))
    
    ## DEMAND REFERENCE (param_Demanda)
    lect = []
    column = sheetDatos['F']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_Demanda= dict(zip(set_REF, lect))
    
    ## FREQUENCY REFERENCE (param_Frecuencia)
    lect = []
    column = sheetDatos['G']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_Frecuencia = dict(zip(set_REF, lect))
    
    ## =======================================================================
    ## RACKS (set_RACKS)
    set_RACKS = []
    column = sheetDatos['L']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_RACKS.append(column[x].value)
        
    
    
    ## IF THE RACK HAVE BACKGROUND (param_TipoRack)
    lect = []
    column = sheetDatos['M']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_TipoRack = dict(zip(set_RACKS, lect))
    
    ## IF THE RACK IS IN WALL (param_Pared)
    lect = []
    column = sheetDatos['N']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_Pared = dict(zip(set_RACKS, lect))
    
    ## IF THE RACK IS IN WALL (param_Utilizacion)
    lect = []
    column = sheetDatos['O']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_Utilizacion = dict(zip(set_RACKS, lect))
    
    ## COST OF ALLOCATE IN THE RACK (param_Costo)
    lect = []
    column = sheetDatos['P']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_Costo = dict(zip(set_RACKS, lect))
    
    
    return( set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo )

    
def create_lineal_model_Alloc( model, set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo ):
    ## --------------------- CONJUNTOS ----------------------------
    model.REF = pyo.Set( initialize = set_REF )
    model.RACKS = pyo.Set( initialize = set_RACKS )
    
    ## ---------------------- PARÁMETROS ----------------------------
    model.AnchoCaja = pyo.Param( model.REF, initialize = param_AnchoCaja )
    model.TipoDist = pyo.Param( model.REF, initialize = param_TipoDist )
    model.Espacios = pyo.Param( model.REF, mutable=True, initialize = param_Espacios )
    model.TipoRack = pyo.Param( model.RACKS, initialize = param_TipoRack )
    model.Pared = pyo.Param( model.RACKS, initialize = param_Pared)
    model.Utilizacion = pyo.Param( model.RACKS, initialize = param_Utilizacion )
    
    model.Demanda  = pyo.Param( model.REF, initialize = param_Demanda )
    model.Frecuencia  = pyo.Param( model.REF, initialize = param_Frecuencia )
    model.Costo = pyo.Param( model.RACKS, initialize = param_Costo )
    
    ## ---------------------- VARIABLES ----------------------------
    model.x = pyo.Var( model.REF, model.RACKS, domain = pyo.Binary )
    model.y = pyo.Var( model.RACKS, domain = pyo.Binary )
    
    ## ---------------------- FUNCIÓN OBJETIVO ----------------------------
    def ObjFunc( model ):
        return sum(model.Demanda[ref]*model.Frecuencia[ref]*model.Costo[rack]*model.x[ref,rack] for ref in model.REF for rack in model.RACKS )
    
    model.FO = pyo.Objective( rule = ObjFunc )
    
    ## ---------------------- RESTRICCIONES ----------------------------
    def r1( model, ref ):
        return sum( model.x[ref,rack] for rack in model.RACKS ) == 1
    
    model.r1 = pyo.Constraint( model.REF, rule = r1 )
    
    def r2( model, rack ):
        return sum( model.AnchoCaja[ref]*model.x[ref,rack] for ref in model.REF ) <= 250*(1-model.y[rack]) + 750*(model.y[rack])
    
    model.r2 = pyo.Constraint( model.RACKS, rule = r2)
    
    def r3( model, ref, rack):
        return model.x[ref,rack] <= (model.y[rack]*model.TipoDist[ref]) + ((1-model.y[rack])*(1-model.TipoDist[ref]))
    
    model.r3 = pyo.Constraint( model.REF, model.RACKS, rule = r3 )
    
    def r6( model, ref, rack ):
        return model.Espacios[ref]*model.x[ref,rack] <= model.TipoRack[rack]
    
    model.r6 = pyo.Constraint( model.REF, model.RACKS, rule = r6 )
    
    def r8( model, rack ):
        return model.y[rack] <= 1 - model.Pared[rack]
    
    model.r8  = pyo.Constraint( model.RACKS, rule = r8 )
     
    def r9( model, ref, rack ):
        return model.x[ref,rack] <= model.Utilizacion[rack]
    
    model.r9 = pyo.Constraint( model.REF, model.RACKS, rule = r9 )





def principal( argv ):
    
    ## READ DATA OPENPYXL
    set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo = read_data_XLSX_Alloc()
    
    ## DECLARE SOLVER, CREATE A PYOMO ABSTRACT MODEL, DECLARE LINEAR MODEL
    opt = SolverFactory('cplex', executable="C:\\Program Files\\IBM\\ILOG\\CPLEX_Studio128\\cplex\\bin\\x64_win64\\cplex")
    model = pyo.AbstractModel()
    create_lineal_model_Alloc( model, set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo )
    
    ## CREATE AN INSTANCE OF THE MODEL, SOLVE PRINT RESULTS BY CONSOLE
    instance = model.create_instance()
    opt.options['timelimit'] = 40
    opt.solve(instance, tee = False)
    #instance.display()
    
    ## PRINT RESULTS IN CONSOLE
    print_results_in_console_Alloc( instance )
    
    ## CREATE EXCEL FILE AND PRINT RESULTS WITH 'XLSXWRITER'
    workbook = xlsxwriter.Workbook('Results_Allocation.xlsx')
    print_results_XLSX_Alloc( workbook, instance )
    


def print_results_XLSX_Alloc( workbook, instance ):
    ## -------------------------- SAVE RESULTS IN EXCEL FILE -------------------------------------

    ## STYLES
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center'})
        
    ## CREATE SHEET
    worksheet = workbook.add_worksheet('Res. REFERENCES')
    worksheet.set_column(0, 0, 16)
    worksheet.set_column(1, 1, 11) ## Width: set_column(first_col, last_col, width, cell_format, options)
    worksheet.set_column(2, 2, 10)
    worksheet.set_column(3, 3, 9.5)
    
    ## PRINT RESULTS
    worksheet.merge_range(0, 0, 0, 3, "SOLUCIÓN ASIGNACIÓN REFERENCIA", merge_format)
    worksheet.merge_range(1, 0, 1, 3, "", merge_format)
    worksheet.write(2, 0, "Función Objetivo: ", merge_format)
    worksheet.merge_range(2, 1, 2, 3, pyo.value(instance.FO), cell_format)
    worksheet.merge_range(3, 0, 3, 3, "", merge_format)
    worksheet.merge_range(4, 0, 4, 3, "Asignación", merge_format)
    
    row = 5
    for ref in instance.REF:
        for rack in instance.RACKS:
            if ( instance.x[ref,rack] == 1 ):
                worksheet.write(row, 0, "La referencia", cell_format)
                worksheet.write(row, 1, ref, cell_format)
                worksheet.write(row, 2, "en el rack", cell_format)
                worksheet.write(row, 3, rack, cell_format)
            
        
        row = row + 1
    
    
    
    
    ## CREATE SHEET 2
    worksheet2 = workbook.add_worksheet('Res. RACKS')
    worksheet2.set_column(0, 2, 16) ## Width: set_column(first_col, last_col, width, cell_format, options)
    
    
    ## PRINT RESULTS
    worksheet2.merge_range(0, 0, 0, 2, "SOLUCIÓN ASIGNACIÓN RACKS", merge_format)
    worksheet2.write(1, 0, "Función Objetivo: ", merge_format)
    worksheet2.merge_range(1, 1, 1, 2, pyo.value(instance.FO), cell_format)
    worksheet2.merge_range(2, 0, 2, 2, "", merge_format)
    worksheet2.merge_range(3, 0, 3, 2, "Asignación", merge_format)
    worksheet2.write(4, 0, "RACK", merge_format)
    worksheet2.write(4, 1, "REFERENCIA", merge_format)
    worksheet2.write(4, 2, "¿HILERAS?", merge_format)
    
    row = 6
    for rack in instance.RACKS:
        ind = False
        for ref in instance.REF:
            if ( instance.x[ref,rack] == 1 ):
                worksheet2.write(row, 0, rack, cell_format)
                worksheet2.write(row, 1, ref, cell_format)
                if ( instance.y[rack] == 1 ):
                    worksheet2.write(row, 2, "si", cell_format)
                else:
                    worksheet2.write(row, 2, "no", cell_format)
                
                row = row + 1
                ind = True
            
        
        if ind == True:
            row = row + 1
        
    
    
    ## CLOSE THE BOOK
    workbook.close()



def print_results_in_console_Alloc( instance ):
    print("\n\n")
    print("SOLUCIÓN DEL EJERCICIO")
    print("--------------------------")
    print("\n")
    print("Función Objetivo: ",pyo.value(instance.FO))
    print("\n")
    for ref in instance.REF:
        for rack in instance.RACKS:
            if ( instance.x[ref,rack] == 1 ):
                print("La referencia ", ref ," en el rack", rack)
            
        
    


if __name__ == "__main__":
    principal( sys.argv )
