import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import openpyxl
from openpyxl.utils import get_column_letter
import copy
import xlsxwriter
import sys


def lecturaDatos():
    ## --------------------- LEER ARCHIVO EXCEL CON OPENPYXL -----------------------------------
    ## Abrir el archivo y guardar en la variable archivo
    archivo = openpyxl.load_workbook('Data_Picking.xlsx', data_only = True)

    ## READ SHEETS OF FILE AND SELECT THE THE FIRST SHEET
    sheets = archivo.sheetnames
    sheetDatos = archivo[sheets[0]]

    ## NODES (set_Nodos)
    set_Nodos = []
    column = sheetDatos['A']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Nodos.append(column[x].value)
        
    

    ## ORDERS (set_Ords)
    set_Ords = []
    column = sheetDatos['B']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Ords.append(column[x].value)
        
    


    ## REFERENCES (set_Referencias)
    set_Referencias = []
    column = sheetDatos['C']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Referencias.append(column[x].value)
        
    

    ## LOCATION OF EACH REFERENCE (param_NOD_REF)
    lect = []
    column = sheetDatos['F']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        
    
    param_NOD_REF = dict(zip(set_Referencias, lect))


    ## Ordenes y Variable auxiliar (set_Ordenes y set_R)
    lect = []
    for i in range(len(set_Ords)):
        column = sheetDatos[get_column_letter( 8 + i )]
        orden = []
        for x in range(len(column)): 
            if (type(column[x].value) != str and column[x].value != None):
                orden.append(column[x].value)
            
        
        lect.append(orden)
    
    set_Ordenes = dict(zip(set_Ords, lect))

    lectR = copy.deepcopy(lect) ## Esto es demencial
    for x in lectR:
        x.remove(0)
    
    set_R = dict(zip(set_Ords, lectR))


    ## DISTANCE MATRIX
    param_Distancia = []
    sheetDistancias = archivo[sheets[1]]
    for i in  range(sheetDistancias.min_row + 1,sheetDistancias.max_row + 1):
        y = []
        for j in range(sheetDistancias.min_column + 1, sheetDistancias.max_column + 1):
            y.append(sheetDistancias.cell(row = i, column = j).value)
        
        param_Distancia.append(y)
    
    ## -------------------------------------------------------------------------------
    
    return( set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia )


def modeloLineal(model, set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia):
    

    ## ------------------- SETS --------------------------------
    model.NODOS = pyo.Set( initialize = set_Nodos )
    model.O = pyo.Set( initialize = set_Ords )
    model.REF = pyo.Set( initialize = set_Referencias )
    def setORDENES(model, i):
        return(list(set_Ordenes[i]))
    
    model.ORDENES = pyo.Set(model.O, initialize = setORDENES)
    def setR(model, i):
        return(list(set_R[i]))
    
    model.R = pyo.Set(model.O, initialize = set_R)



    def junte(model):
        return((o,i,j) for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o] )
    
    model.OROR = pyo.Set(dimen =3, initialize = junte )

    def junte2(model):
        return((o,i) for o in model.O for i in model.ORDENES[o] )
    
    model.OX = pyo.Set(dimen = 2, initialize = junte2 )

    def junte3(model):
        return((o,i,j) for o in model.O for i in model.R[o] for j in model.R[o] )
    
    model.ORR = pyo.Set(dimen = 3, initialize = junte3 )

    def junte4(model):
        return((o,i) for o in model.O for i in model.R[o] )
    
    model.OR  = pyo.Set(dimen = 2, initialize = junte4 )


    ## --------------------- PARAMETERS -------------------------------------
    def paramDistancia(model, i, j):
        return(param_Distancia[i][j])
    
    model.Distancia = pyo.Param(model.NODOS,model.NODOS, initialize = paramDistancia)
    
    def paramNodRef(model, i):
        return(param_NOD_REF[i])
    
    model.Nod_Ref = pyo.Param(model.REF, initialize = paramNodRef) 

    def paramDistR(model, i ,j):
        return model.Distancia[model.Nod_Ref[i],model.Nod_Ref[j]]
    
    model.Distancia_R = pyo.Param(model.REF, model.REF, initialize = paramDistR, mutable = True)

    ## ----------------------- VARIABLES ---------------------------------------
    model.x = pyo.Var(model.OROR, domain = pyo.Binary)
    model.aux = pyo.Var(model.OR)
    model.Dist_ORD = pyo.Var(model.O)


    ## -------------------------- OBJECTIVE FUNCTION ------------------------------------
    def ObjFunc(model):
        return sum(model.Distancia_R[i,j]*model.x[o,i,j] for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o])
    
    model.FO = pyo.Objective(rule = ObjFunc)


    ## --------------------------- RESTRICTIONS ----------------------------------------------
    def r1(model, o, i):
        return sum( model.x[o,i,j] for j in model.ORDENES[o]) == 1
    
    model.r1 = pyo.Constraint( model.OX, rule = r1 )

    def r2(model, o, j):
        return sum( model.x[o,i,j] for i in model.ORDENES[o]) == 1
    
    model.r2 = pyo.Constraint( model.OX,  rule = r2 )

    def r3(model, o, i):
        return model.x[o,i,i] == 0
    
    model.r3 = pyo.Constraint( model.OX,  rule = r3 )

    def r4(model, o, i, j):
        if ( i != j ):
            return model.aux[o,i] - model.aux[o,j] + len(model.R[o])*model.x[o,i,j] <= len(model.R[o]) - 1 
        return pyo.Constraint.Skip
    
    model.r4 = pyo.Constraint( model.ORR,  rule = r4 )

    def r5(model, o):
        return model.Dist_ORD[o] == sum( model.x[o,i,j]*model.Distancia_R[i,j] for i in model.ORDENES[o] for j in model.ORDENES[o] )
    
    model.r5 = pyo.Constraint( model.O, rule = r5)




def imprimirResultadosXLSX(instance):
    ## --------------------- GUARDAR LOS RESULTADOS EN ARCHIVO EXCEL -----------------------------
    
    # Crear libro
    workbook = xlsxwriter.Workbook('Resultados.xlsx')

    ## Formatos
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center'})
        
    ## Crear hoja
    worksheet = workbook.add_worksheet('Resultados')
    worksheet.set_column(0, 0, 20)

    #worksheet.merge_range('A1:D1', "SOLUCIÓN DEL EJERCICIO", merge_format)
    worksheet.merge_range(0, 0, 0, 3, "SOLUCIÓN DEL EJERCICIO", merge_format)

    worksheet.write(1, 0, "Distancia Total: ", merge_format)
    worksheet.write(1, 1, pyo.value(instance.FO), cell_format)

    row=3
    col=0
    for o in instance.O:
        worksheet.merge_range(row, col, row, 3, "Orden "+ str(o), merge_format)
        row += 1

        worksheet.write(row, col,"Distancia Recorrida: ", cell_format)
        worksheet.write(row, col + 1,instance.Dist_ORD[o].value, cell_format)
        row += 1
        
        worksheet.write(row, col,"Ruta: ", cell_format)
        row += 1
        
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    worksheet.write(row, col + 0,"Del nodo ", cell_format)
                    worksheet.write(row, col + 1,i, cell_format)
                    worksheet.write(row, col + 2," al nodo", cell_format)
                    worksheet.write(row, col + 3,j, cell_format)
                
            
            row += 1
        
        row += 1
    
        
        
    # Hay que incluir workbook.close()
    workbook.close()
    ##  -------------------------------------------------------------------------------------





## Impresión de Resultados
def imprimirResultadosConsola(instance):
    print("\n\n")
    print("SOLUCIÓN DEL EJERCICIO")
    print("--------------------------")
    print("\n")
    print("Distancia Total: ", pyo.value(instance.FO))
    print("\n")
    for o in instance.O:
        print("--------- Orden ",o,"----------")
        print("Distancia Recorrida: ",instance.Dist_ORD[o].value)
        print("Ruta: ")
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    print("Del nodo ",i," al nodo",j)
                
            
        
    




def principal( argv ):
    ## Lectura de datos con OPENPYXL
    set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia = lecturaDatos()
    
    
    ## Crear Modelo Lineal con PYOMO, resolverlo con GLPK y guardarlo en variable model
    opt = SolverFactory('cplex', executable="C:\\Program Files\\IBM\\ILOG\\CPLEX_Studio1210\\cplex\\bin\\x64_win64\\cplex")
    #opt = SolverFactory('glpk')
    model = pyo.AbstractModel()
    modeloLineal(model, set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia)
    
    ## Crear una instancia del modelo y ejecutarla
    instance = model.create_instance() 
    opt.options['timelimit'] = 120
    opt.solve(instance, tee = False)
    #instance.display()
    
    
    ## Imprimir resultado por consola
    imprimirResultadosConsola(instance)
    
    ## Crear archivo e imprimir resultados xlsxwriter
    imprimirResultadosXLSX(instance)

        
if __name__ == "__main__":
    principal( sys.argv )
